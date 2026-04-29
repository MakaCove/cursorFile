#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将测试用例 Markdown 表格（.md）导出为 Excel 文件（.xlsx）。

支持两类用例表，分别写入不同 sheet：
  - 功能点用例（表头含"模块""ID""标题"）
  - 业务链路用例（表头含"链路ID""链路名称""涉及模块"）

用法:
  python md_testcases_to_excel.py <testcases.md> [-o output.xlsx]

依赖: pip install openpyxl
"""
import sys
import argparse
from pathlib import Path
import re
from typing import Optional

# 功能点用例表头
FUNCTIONAL_HEADERS = ["模块", "ID", "标题", "优先级", "类型", "前置条件", "步骤", "测试数据", "预期结果", "备注/覆盖点"]

# 业务链路用例表头
CHAIN_HEADERS = ["链路ID", "链路名称", "涉及模块", "优先级", "类型", "前置条件", "步骤", "测试数据", "预期结果", "备注/覆盖点"]

_RE_STEPS = re.compile(r"(?<!\n)(?<!^)\s*(?=(\[\d+\]|\d+\.))")
_RE_EXPECT_BRACKET = re.compile(r"(?<!\n)(?<!^)\s*(?=\[[^\]]*\d+[^\]]*\])")
_RE_EXPECT_PAREN = re.compile(r"(?<!\n)(?<!^)\s*(?=\d+[）\)])")
_RE_DATA_SEP = re.compile(r"(?<!\n)(?<!^)\s*[;；、]\s*")

_RE_MD_TABLE_ROW = re.compile(r"^\s*\|.*\|\s*$")
_RE_MD_TABLE_SEP = re.compile(r"^\s*\|\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|\s*$")


def normalize_cell_text(text: str, header: str) -> str:
    """
    将单元格文本规范化为"单元格内换行"的形式。
    - 对「步骤」: 支持 '1. ... 2. ...' 和 '[1] ... [2] ...' 格式自动拆成多行
    - 对「预期结果」: 支持 '1）... 2）...' / '1)... 2)...' / '[1] ... [2] ...' 格式自动拆成多行
    其它字段保持原样（仅 trim）。
    """
    if text is None:
        return ""
    s = str(text).strip()
    if not s or "\n" in s:
        return s

    if header == "步骤":
        return _RE_STEPS.sub("\n", s)
    if header == "预期结果":
        s = _RE_EXPECT_BRACKET.sub("\n", s)
        s = _RE_EXPECT_PAREN.sub("\n", s)
        return s
    if header == "测试数据":
        return _RE_DATA_SEP.sub("\n", s)
    return s


def get_cell_value(row: dict, header: str) -> str:
    """从一行数据中取与表头对应的值。"""
    val = row.get(header, "")
    return normalize_cell_text(val, header)


def _split_md_row(line: str) -> list[str]:
    """
    Split a markdown table row into cells.
    Supports escaped pipe: \\|
    """
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]

    cells: list[str] = []
    cur: list[str] = []
    escape = False
    for ch in s:
        if escape:
            cur.append(ch)
            escape = False
            continue
        if ch == "\\":
            escape = True
            continue
        if ch == "|":
            cells.append("".join(cur).strip())
            cur = []
            continue
        cur.append(ch)
    cells.append("".join(cur).strip())
    return cells


def _parse_table(lines: list[str], start: int, headers: list[str]) -> tuple[Optional[list[dict]], int]:
    """
    尝试从 start 行开始解析一个表格。
    如果表头匹配 headers，返回 (rows, consumed_lines)。
    否则返回 (None, 0)。
    """
    if start >= len(lines):
        return None, 0

    for i in range(start, len(lines)):
        line = lines[i]
        if not _RE_MD_TABLE_ROW.match(line):
            continue
        header_cells = _split_md_row(line)
        if not header_cells:
            continue
        if all(h in header_cells for h in headers):
            if i + 1 >= len(lines) or not _RE_MD_TABLE_SEP.match(lines[i + 1]):
                continue

            col_index = {name: header_cells.index(name) for name in headers}
            rows: list[dict] = []
            for j in range(i + 2, len(lines)):
                row_line = lines[j]
                if not row_line.strip():
                    continue
                if not _RE_MD_TABLE_ROW.match(row_line):
                    return rows, j - start
                row_cells = _split_md_row(row_line)
                if len(row_cells) < len(header_cells):
                    row_cells += [""] * (len(header_cells) - len(row_cells))
                row = {h: row_cells[col_index[h]] for h in headers}
                rows.append(row)
            if rows:
                return rows, j - start + 1
            return None, 0
    return None, 0


def load_tables_from_md(md_path: Path) -> tuple[list[dict], list[dict]]:
    """
    从 Markdown 中分别解析功能点用例表和业务链路用例表。
    返回 (functional_rows, chain_rows)。
    """
    text = md_path.read_text(encoding="utf-8")
    lines = text.splitlines()

    functional_rows: list[dict] = []
    chain_rows: list[dict] = []

    pos = 0
    while pos < len(lines):
        # 尝试匹配功能点表
        rows, consumed = _parse_table(lines, pos, FUNCTIONAL_HEADERS)
        if rows:
            functional_rows = rows
            pos += max(consumed, 1)
            continue

        # 尝试匹配链路表
        rows, consumed = _parse_table(lines, pos, CHAIN_HEADERS)
        if rows:
            chain_rows = rows
            pos += max(consumed, 1)
            continue

        pos += 1

    if not functional_rows and not chain_rows:
        raise ValueError("未在 Markdown 中找到包含完整表头的测试用例表格。")

    return functional_rows, chain_rows


def write_sheet(ws, headers: list[str], rows: list[dict]) -> None:
    """将表头和行数据写入 worksheet。"""
    from openpyxl.styles import Font, Alignment
    import openpyxl

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row in enumerate(rows, start=2):
        for c, header in enumerate(headers, start=1):
            val = get_cell_value(row, header)
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16


def main() -> None:
    parser = argparse.ArgumentParser(description="导出测试用例 Markdown 表格为 Excel")
    parser.add_argument("md_file", help="测试用例 Markdown 文件路径（.md，表格格式）")
    parser.add_argument("-o", "--output", default=None, help="输出 .xlsx 路径，默认与 MD 同目录同主名")
    args = parser.parse_args()

    md_path = Path(args.md_file)
    if not md_path.exists():
        print(f"错误: 文件不存在 {md_path}", file=sys.stderr)
        sys.exit(1)
    if md_path.suffix.lower() != ".md":
        print("错误: 仅支持 .md 输入", file=sys.stderr)
        sys.exit(1)

    try:
        functional_rows, chain_rows = load_tables_from_md(md_path)
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("错误: 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()

    if functional_rows and chain_rows:
        # Two sheets
        ws1 = wb.active
        ws1.title = "功能点用例"
        write_sheet(ws1, FUNCTIONAL_HEADERS, functional_rows)

        ws2 = wb.create_sheet("业务链路用例")
        write_sheet(ws2, CHAIN_HEADERS, chain_rows)
    elif functional_rows:
        ws = wb.active
        ws.title = "功能点用例"
        write_sheet(ws, FUNCTIONAL_HEADERS, functional_rows)
    elif chain_rows:
        ws = wb.active
        ws.title = "业务链路用例"
        write_sheet(ws, CHAIN_HEADERS, chain_rows)

    out_path = args.output
    if not out_path:
        out_path = md_path.with_suffix(".xlsx")
    else:
        out_path = Path(out_path)
    wb.save(out_path)
    sheet_info = []
    if functional_rows:
        sheet_info.append(f"功能点用例({len(functional_rows)}条)")
    if chain_rows:
        sheet_info.append(f"业务链路用例({len(chain_rows)}条)")
    print(f"已导出: {out_path}  [{', '.join(sheet_info)}]")


if __name__ == "__main__":
    main()
