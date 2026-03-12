#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将测试用例 Excel（.xlsx）导出为 Markdown（.md）表格。

用法:
- python xlsx_testcases_to_md.py <testcases.xlsx> [-o output.md] [--sheet all|active|SheetName]

说明:
- 自动识别包含用例表头的 sheet（表头需包含 HEADERS 中的大部分列）。
- 输出的 Markdown 仅包含用例表格（可被 md_testcases_to_xmind.py / md_testcases_to_excel.py 复用）。
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


HEADERS = ["模块", "ID", "标题", "优先级", "类型", "前置条件", "步骤", "测试数据", "预期结果", "备注/覆盖点"]


def _norm(s: str) -> str:
    return (s or "").strip()


def _cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v)
    # Keep markdown table row as single line.
    # Convert cell-internal newlines to <br> for readability.
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    return s.strip()


def _escape_md_cell(s: str) -> str:
    # escape pipe to keep table structure
    return s.replace("|", "\\|")


def _find_header_row(rows: Iterable[list[str]]) -> tuple[int, list[str]] | None:
    """
    Return (row_index_1_based, header_cells) if a header row is found.
    """
    for idx, r in enumerate(rows, start=1):
        cells = [_norm(x) for x in r]
        if not any(cells):
            continue
        hits = sum(1 for h in HEADERS if h in cells)
        if hits >= 6 and "ID" in cells and "标题" in cells:
            return idx, cells
    return None


def _read_sheet_as_rows(ws, max_scan_rows: int = 50) -> list[list[str]]:
    out: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        out.append([_cell_str(v) for v in row])
        if i >= max_scan_rows:
            break
    return out


def extract_testcases_from_sheet(ws) -> tuple[list[str], list[dict]] | None:
    """
    Returns (sheet_title, rows) where rows is list[dict] keyed by HEADERS.
    """
    preview = _read_sheet_as_rows(ws, max_scan_rows=80)
    found = _find_header_row(preview)
    if not found:
        return None
    header_row_idx, header_cells = found

    # Map required headers to column index
    col_index = {}
    for h in HEADERS:
        if h in header_cells:
            col_index[h] = header_cells.index(h) + 1  # 1-based

    # Require at least ID & 标题
    if "ID" not in col_index or "标题" not in col_index:
        return None

    rows: list[dict] = []
    # iterate from the next row after header
    for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # stop if entire row empty
        if r is None:
            continue
        if all(v is None or str(v).strip() == "" for v in r):
            # allow trailing empty rows; continue scanning a bit is expensive; break here
            continue
        row_dict = {}
        # If ID is empty, skip
        id_val = _cell_str(r[col_index["ID"] - 1]) if col_index["ID"] - 1 < len(r) else ""
        title_val = _cell_str(r[col_index["标题"] - 1]) if col_index["标题"] - 1 < len(r) else ""
        if not id_val and not title_val:
            continue
        for h in HEADERS:
            if h in col_index and col_index[h] - 1 < len(r):
                row_dict[h] = _cell_str(r[col_index[h] - 1])
            else:
                row_dict[h] = ""
        rows.append(row_dict)

    if not rows:
        return None
    return ws.title, rows


def to_markdown(sheets_data: list[tuple[str, list[dict]]], title: str) -> str:
    lines: list[str] = []
    lines.append(f"# {title}")
    lines.append("")

    for sheet_name, rows in sheets_data:
        if len(sheets_data) > 1:
            lines.append(f"## {sheet_name}")
            lines.append("")

        lines.append("| " + " | ".join(HEADERS) + " |")
        lines.append("|" + "|".join(["------"] * len(HEADERS)) + "|")

        for r in rows:
            cells = [_escape_md_cell(_cell_str(r.get(h, "")) or "—") for h in HEADERS]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


@dataclass(frozen=True)
class Args:
    xlsx_file: str
    output: str | None
    sheet: str


def main() -> None:
    parser = argparse.ArgumentParser(description="将测试用例 Excel 导出为 Markdown 表格")
    parser.add_argument("xlsx_file", help="测试用例 Excel 文件路径（.xlsx）")
    parser.add_argument("-o", "--output", default=None, help="输出 .md 路径，默认与 xlsx 同目录同主名")
    parser.add_argument("--sheet", default="all", help="导出 sheet：all | active | <SheetName>")
    ns = parser.parse_args()
    args = Args(xlsx_file=ns.xlsx_file, output=ns.output, sheet=ns.sheet)

    xlsx_path = Path(args.xlsx_file)
    if not xlsx_path.exists():
        print(f"错误: 文件不存在 {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    if xlsx_path.suffix.lower() != ".xlsx":
        print("错误: 仅支持 .xlsx 输入", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("错误: 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    target_sheets = []
    if args.sheet == "active":
        target_sheets = [wb.active]
    elif args.sheet == "all":
        target_sheets = [wb[s] for s in wb.sheetnames]
    else:
        if args.sheet not in wb.sheetnames:
            print(f"错误: 不存在的 sheet：{args.sheet}", file=sys.stderr)
            sys.exit(1)
        target_sheets = [wb[args.sheet]]

    extracted: list[tuple[str, list[dict]]] = []
    for ws in target_sheets:
        got = extract_testcases_from_sheet(ws)
        if got:
            extracted.append((got[0], got[1]))

    if not extracted:
        print("错误: 未在 Excel 中找到包含测试用例表头的 sheet（需要包含：模块/ID/标题/…）", file=sys.stderr)
        sys.exit(1)

    title = xlsx_path.stem
    md = to_markdown(extracted, title=title)

    out = Path(args.output) if args.output else xlsx_path.with_suffix(".md")
    out.write_text(md, encoding="utf-8")
    print(f"已导出: {out}")


if __name__ == "__main__":
    main()

