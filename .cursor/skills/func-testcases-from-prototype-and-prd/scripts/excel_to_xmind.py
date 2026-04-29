#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将测试用例 Excel（.xlsx）直接转为 XMind 脑图（.xmind，Zen/2020+ 兼容）。

每个 Sheet 对应 XMind 中的一个一级分支节点，Sheet 内的用例按表头类型自动分组。
支持可选输出中间 Markdown 文件。

用法:
  python excel_to_xmind.py <testcases.xlsx> [-o output.xmind] [--md] [--group-by 模块|优先级|类型]

说明:
  - 自动识别每个 Sheet 中的测试用例表格（功能点 / 业务链路）。
  - 多个 Sheet → XMind 中多个一级分支。
  - --md：同时生成合并后的 Markdown 文件。
  - 此脚本为按需手动调用，不会在生成测试用例时自动执行。
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import re
import uuid
from typing import Optional


FUNCTIONAL_HEADERS = ["模块", "ID", "标题", "优先级", "类型", "前置条件", "步骤", "测试数据", "预期结果", "备注/覆盖点"]
CHAIN_HEADERS = ["链路ID", "链路名称", "涉及模块", "优先级", "类型", "前置条件", "步骤", "测试数据", "预期结果", "备注/覆盖点"]


def _new_id(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:12]}"


def _norm(s: str) -> str:
    return (s or "").strip()


def _cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    return s.strip()


def _topic(title: str, children: list[dict] | None = None) -> dict:
    t = {"id": _new_id("topic"), "class": "topic", "title": title}
    if children:
        t["children"] = {"attached": children}
    return t


def _split_lines(s: str) -> list[str]:
    t = _norm(s)
    if not t:
        return []
    t = t.replace("<br>", "\n")
    return [x.strip() for x in t.splitlines() if x.strip()]


def _find_header_row(rows: list[list[str]]) -> tuple[int, list[str]] | None:
    """在 sheet 行列表中寻找包含足够多表头列的行。返回 (row_index_0based, header_cells)。"""
    for idx, r in enumerate(rows):
        cells = [_norm(x) for x in r]
        if not any(cells):
            continue
        # 功能点头：检查 模块, ID, 标题
        if "模块" in cells and "ID" in cells and "标题" in cells:
            return idx, cells
        # 链路头：检查 链路ID, 链路名称, 涉及模块
        if "链路ID" in cells and "链路名称" in cells and "涉及模块" in cells:
            return idx, cells
        # 宽松匹配：ID + 标题
        if "ID" in cells and "标题" in cells:
            return idx, cells
    return None


def _read_sheet_rows(ws, max_scan: int = 100) -> list[list[str]]:
    out: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        out.append([_cell_str(v) for v in row])
        if i >= max_scan:
            break
    return out


def _detect_header_type(header_cells: list[str]) -> str:
    """返回 'functional' 或 'chain'."""
    if "链路ID" in header_cells or "链路名称" in header_cells or "涉及模块" in header_cells:
        return "chain"
    return "functional"


def _extract_rows(ws, header_row_idx: int, header_cells: list[str]) -> list[dict]:
    """从表头下一行开始提取数据行，直到空行或数据结束。"""
    htype = _detect_header_type(header_cells)
    if htype == "chain":
        target_headers = CHAIN_HEADERS
    else:
        target_headers = FUNCTIONAL_HEADERS

    col_index: dict[str, int] = {}
    for h in target_headers:
        if h in header_cells:
            col_index[h] = header_cells.index(h)

    if ("ID" not in col_index and "链路ID" not in col_index) or "标题" not in col_index:
        return []

    rows: list[dict] = []
    for r in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):  # +2: 1-based → skip header
        if r is None:
            continue
        if all(v is None or str(v).strip() == "" for v in r):
            continue

        row_dict: dict[str, str] = {}
        has_id = False
        for h in target_headers:
            if h in col_index and col_index[h] < len(r):
                val = _cell_str(r[col_index[h]])
                row_dict[h] = val
                if h in ("ID", "链路ID") and val:
                    has_id = True
            else:
                row_dict[h] = ""

        if has_id or _norm(row_dict.get("标题", "")):
            rows.append(row_dict)

    return rows


def _escape_md_cell(s: str) -> str:
    return s.replace("|", "\\|")


def _rows_to_markdown(
    sheets_data: list[tuple[str, str, list[dict]]],  # (sheet_name, table_type, rows)
    title: str,
) -> str:
    """将多个 sheet 的用例行合并为一个 Markdown 文档。"""
    lines: list[str] = []
    lines.append(f"# {title}")
    lines.append("")

    for sheet_name, htype, rows in sheets_data:
        lines.append(f"## {sheet_name}")
        lines.append("")

        if htype == "chain":
            headers = CHAIN_HEADERS
        else:
            headers = FUNCTIONAL_HEADERS

        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join(["------"] * len(headers)) + "|")

        for r in rows:
            cells = [_escape_md_cell(_norm(r.get(h, "")) or "—") for h in headers]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


# ── XMind building ──────────────────────────────────────────────


def _build_case_detail_topics(r: dict, is_chain: bool) -> list[dict]:
    """为单条用例构建详情子节点列表（复用现有 XMind 脚本逻辑）。"""
    if is_chain:
        tid = _norm(r.get("链路ID", ""))
        title = _norm(r.get("链路名称", ""))
        modules_val = _norm(r.get("涉及模块", ""))
        case_title = f"{tid} {title}".strip() if tid or title else "（未命名链路）"
    else:
        tid = _norm(r.get("ID", ""))
        title = _norm(r.get("标题", ""))
        case_title = f"{tid} {title}".strip() if tid or title else "（未命名用例）"

    details: list[dict] = []

    if is_chain:
        modules_val = _norm(r.get("涉及模块", ""))
        if modules_val:
            details.append(_topic(f"涉及模块：{modules_val}"))

    prio = _norm(r.get("优先级", ""))
    typ = _norm(r.get("类型", ""))
    pre = _norm(r.get("前置条件", ""))
    steps = _norm(r.get("步骤", ""))
    data = _norm(r.get("测试数据", ""))
    exp = _norm(r.get("预期结果", ""))
    note = _norm(r.get("备注/覆盖点", ""))

    if prio:
        details.append(_topic(f"优先级：{prio}"))
    if typ:
        details.append(_topic(f"类型：{typ}"))
    if pre:
        details.append(_topic(f"前置：{pre}"))
    if steps:
        step_children = [_topic(x) for x in _split_lines(steps)]
        details.append(_topic("步骤", step_children if step_children else None))
    if data and data != "—":
        details.append(_topic(f"数据：{data}"))
    if exp:
        exp_children = [_topic(x) for x in _split_lines(exp)]
        details.append(_topic("预期", exp_children if exp_children else None))
    if note:
        details.append(_topic(f"备注：{note}"))

    return details, case_title


def _build_sheet_branch(
    sheet_name: str,
    htype: str,
    rows: list[dict],
    group_by: str,
) -> dict:
    """为一个 Sheet 构建 XMind 分支节点。返回 _topic dict。"""
    if htype == "chain":
        # Group by chain type (from 备注/覆盖点)
        groups: dict[str, list[dict]] = {}
        for r in rows:
            note = _norm(r.get("备注/覆盖点", ""))
            chain_type = "（未分类链路）"
            if "主链路" in note:
                chain_type = "主链路"
            elif "分支链路" in note:
                chain_type = "分支链路"
            elif "中断恢复链路" in note or "中断恢复" in note:
                chain_type = "中断恢复链路"
            elif "回退链路" in note:
                chain_type = "回退链路"
            elif "超时中断" in note:
                chain_type = "超时中断链路"

            if chain_type not in groups:
                groups[chain_type] = []
            groups[chain_type].append(r)

        type_order = {"主链路": 0, "分支链路": 1, "中断恢复链路": 2, "超时中断链路": 3, "回退链路": 4}
        group_topics: list[dict] = []
        for g in sorted(groups.keys(), key=lambda k: type_order.get(k, 9)):
            case_topics: list[dict] = []
            for r in groups[g]:
                details, case_title = _build_case_detail_topics(r, is_chain=True)
                case_topics.append(_topic(case_title, details if details else None))
            group_topics.append(_topic(g, case_topics))
        return _topic(sheet_name, group_topics)
    else:
        # Group functional by 模块/优先级/类型
        groups: dict[str, list[dict]] = {}
        module_order: list[str] = []
        for r in rows:
            if group_by == "模块":
                k = _norm(r.get("模块", "")) or "（未填模块）"
            elif group_by == "类型":
                k = _norm(r.get("类型", "")) or "（未填类型）"
            else:
                k = _norm(r.get("优先级", "")) or "（未填优先级）"

            if k not in groups:
                groups[k] = []
                if group_by == "模块":
                    module_order.append(k)
            groups[k].append(r)

        group_topics: list[dict] = []
        if group_by == "优先级":
            order = {"P0": 0, "P1": 1, "P2": 2}
            group_iter = sorted(groups.keys(), key=lambda k: (order.get(k, 9), k))
        elif group_by == "模块":
            group_iter = module_order
        else:
            group_iter = sorted(groups.keys())

        for g in group_iter:
            case_topics: list[dict] = []
            for r in groups[g]:
                details, case_title = _build_case_detail_topics(r, is_chain=False)
                case_topics.append(_topic(case_title, details if details else None))
            group_topics.append(_topic(g, case_topics))
        return _topic(sheet_name, group_topics)


def build_xmind_from_sheets(
    sheets_data: list[tuple[str, str, list[dict]]],
    root_title: str,
    group_by: str,
) -> list[dict]:
    """从多个 Sheet 的数据构建 XMind content JSON。"""
    branch_topics: list[dict] = []
    for sheet_name, htype, rows in sheets_data:
        if not rows:
            continue
        branch = _build_sheet_branch(sheet_name, htype, rows, group_by)
        branch_topics.append(branch)

    sheet_id = _new_id("sheet")
    root = _topic(root_title, branch_topics if branch_topics else None)
    root.setdefault("structureClass", "org.xmind.ui.logic.right")

    sheet = {
        "id": sheet_id,
        "class": "sheet",
        "title": "Sheet 1",
        "rootTopic": root,
    }
    return [sheet]


def write_xmind(output_path: Path, content: list[dict]) -> None:
    """Write a minimal .xmind (zip) for XMind Zen/2020+."""
    created = datetime.now(timezone.utc).isoformat()
    active_sheet_id = content[0]["id"] if content and isinstance(content, list) else None
    metadata = {
        "creator": {
            "name": "excel_to_xmind.py",
            "version": "1.0",
        },
        "activeSheetId": active_sheet_id,
        "created": created,
        "modified": created,
    }
    manifest = {
        "file-entries": {
            "content.json": {},
            "metadata.json": {},
            "manifest.json": {},
        }
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))
        z.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        z.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))


@dataclass(frozen=True)
class Args:
    xlsx_file: str
    output: str | None
    group_by: str
    save_md: bool


def main() -> None:
    parser = argparse.ArgumentParser(
        description="将测试用例 Excel 转为 XMind 脑图（每个 Sheet → 一个一级分支）"
    )
    parser.add_argument("xlsx_file", help="测试用例 Excel 文件路径（.xlsx）")
    parser.add_argument("-o", "--output", default=None, help="输出 .xmind 路径，默认与 xlsx 同目录同主名")
    parser.add_argument(
        "--group-by",
        default="模块",
        choices=["优先级", "类型", "模块"],
        help="功能点用例分组维度（默认：模块）。业务链路用例始终按链路类型分组。",
    )
    parser.add_argument(
        "--md",
        action="store_true",
        dest="save_md",
        help="同时生成合并后的 Markdown 文件（.md）",
    )
    ns = parser.parse_args()
    args = Args(xlsx_file=ns.xlsx_file, output=ns.output, group_by=ns.group_by, save_md=ns.save_md)

    xlsx_path = Path(args.xlsx_file)
    if not xlsx_path.exists():
        print(f"错误: 文件不存在 {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    if xlsx_path.suffix.lower() not in (".xlsx", ".xlsm"):
        print("错误: 仅支持 .xlsx / .xlsm 输入", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("错误: 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 遍历所有 sheet，提取测试用例表格
    sheets_data: list[tuple[str, str, list[dict]]] = []
    total_rows = 0

    for ws in wb.worksheets:
        preview = _read_sheet_rows(ws, max_scan=100)
        found = _find_header_row(preview)
        if not found:
            print(f"跳过 Sheet「{ws.title}」：未找到测试用例表头")
            continue

        header_row_idx, header_cells = found
        htype = _detect_header_type(header_cells)
        rows = _extract_rows(ws, header_row_idx, header_cells)

        if rows:
            sheets_data.append((ws.title, htype, rows))
            total_rows += len(rows)
            type_label = "功能点" if htype == "functional" else "业务链路"
            print(f"Sheet「{ws.title}」→ {type_label}用例 {len(rows)} 条")
        else:
            print(f"跳过 Sheet「{ws.title}」：表头下无有效数据行")

    if not sheets_data:
        print("错误: 所有 Sheet 中均未找到有效的测试用例表格。", file=sys.stderr)
        sys.exit(1)

    root_title = xlsx_path.stem

    # 可选：输出合并 Markdown
    if args.save_md:
        md_content = _rows_to_markdown(sheets_data, title=root_title)
        md_path = xlsx_path.with_suffix(".md")
        md_path.write_text(md_content, encoding="utf-8")
        print(f"已导出 MD: {md_path}")

    # 构建 XMind
    content = build_xmind_from_sheets(sheets_data, root_title=root_title, group_by=args.group_by)

    out = Path(args.output) if args.output else xlsx_path.with_suffix(".xmind")
    write_xmind(out, content)
    print(f"已导出: {out}  [共 {len(sheets_data)} 个Sheet, {total_rows} 条用例]")


if __name__ == "__main__":
    main()
